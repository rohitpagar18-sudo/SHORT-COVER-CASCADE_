"""Phase 5D — Paper Trades + Paper Dashboard + Echoes sheets.

Three new sheets appended to the existing quarterly workbook:

  1. ``Paper Trades``      — one row per episode representative
                              (TAKEN + SKIPPED). Coloured by outcome.
  2. ``Paper Dashboard``   — KPI cards built with Excel formulas
                              over Paper Trades. Manual columns win.
  3. ``Echoes (diagnostic)`` — every non-representative (echo)
                              re-fire. Hidden by default; collapse-
                              ratio data lives here.

The sheets read from a ``paper_trades.jsonl`` produced by
``src.paper.engine.run_paper_engine``. The dashboard builder calls
the engine once per workbook refresh, so the on-disk JSONL and the
sheets stay in lock-step. Manual overrides come from the user's
``paper_overrides.csv`` and ALWAYS win over the auto values.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

import pandas as pd
from loguru import logger
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

IST = ZoneInfo("Asia/Kolkata")


# ---------------------------------------------------------------------------
# Paper-specific palette (use judgment, make it readable)
# ---------------------------------------------------------------------------

PAPER_HEADER_FILL = PatternFill("solid", fgColor="2E7D32")  # deep green
PAPER_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
PAPER_BODY_FONT = Font(name="Calibri", size=11)
PAPER_TITLE_FONT = Font(name="Calibri", bold=True, color="2E7D32", size=16)
PAPER_SUB_FONT = Font(name="Calibri", italic=True, color="595959", size=10)

# Per-row fill by paper outcome (D5 spec).
PAPER_OUTCOME_FILLS = {
    "TP2_HIT":     PatternFill("solid", fgColor="6FBF73"),  # deep green
    "TP1_BE":      PatternFill("solid", fgColor="C8E6C9"),  # light green
    "TP1_HIT":     PatternFill("solid", fgColor="C8E6C9"),  # light green
    "SL_HIT":      PatternFill("solid", fgColor="EF9A9A"),  # red
    "HARD_EXIT":   PatternFill("solid", fgColor="EF9A9A"),  # red
    "OPEN_SQOFF":  PatternFill("solid", fgColor="FFD180"),  # amber
    "NO_DATA":     PatternFill("solid", fgColor="EEEEEE"),  # hatched grey
}

PAPER_SKIPPED_FILL = PatternFill("solid", fgColor="E0E0E0")  # grey

OUTCOME_CHIP = {
    "TP2_HIT":    "🟢 TP2",
    "TP1_BE":     "🟡 TP1→BE",
    "TP1_HIT":    "🟢 TP1",
    "SL_HIT":     "🔴 SL",
    "HARD_EXIT":  "🔴 Hard",
    "OPEN_SQOFF": "⏹ SqOff",
    "NO_DATA":    "· N/A",
}

MANUAL_OUTCOME_CHOICES = "TP2,TP1_BE,SL,Whipsaw,No-fill,Skipped,Open,Indeterminate"
MANUAL_DECISION_CHOICES = "TAKEN,SKIPPED,ECHO"

PAPER_SHEET_BANNER = (
    "PAPER — reconstructed from cached OHLC, not real fills. "
    "First-alert-only."
)

# Row where the Paper Trades column headers live (data starts at +1).
# Paired constants — update both if the help/banner block changes size.
PAPER_TRADES_HEADER_ROW = 10
PAPER_TRADES_DATA_START = PAPER_TRADES_HEADER_ROW + 1


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _set_paper_headers(ws: Worksheet, headers: Iterable[str], row: int = 1) -> None:
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = PAPER_HEADER_FONT
        c.fill = PAPER_HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _autofit(ws: Worksheet, df: pd.DataFrame, header_row: int = 1) -> None:
    for col_idx, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].head(200).tolist()]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(12, max_len + 2), 38
        )


def _outcome_chip(outcome: str | None, decision: str | None) -> str:
    if decision and str(decision).upper() == "SKIPPED":
        return "· skip"
    if outcome is None:
        return ""
    return OUTCOME_CHIP.get(str(outcome).upper(), str(outcome))


# ---------------------------------------------------------------------------
# Paper Trades sheet
# ---------------------------------------------------------------------------


PAPER_TRADE_COLUMNS = [
    # Trade identifier (when + what)
    "date", "candle_timestamp", "symbol", "strike", "option_type",
    # Levels
    "entry", "paper_pnl", "outcome", "result_chip","exit_price", "sl", "tp1", "tp2","exit_time", "lots","is_expiry_day",
    "relation","decision","lot_size", "expiry",
    "decision_reason", "slot",
    # Detailed metrics
    "realized_R", "mfe", "mae", "mfe_R", "mae_R",
    "max_drawdown_R", "intrabar_ambiguous",
    # Identity / diagnostics
    "alert_id", "episode_id", "bot_remark", "bot_tags", "triggered_caps",
    # Manual override columns — END (user fills these to override auto)
    "manual_decision", "manual_reason", "manual_outcome",
    "manual_exit", "user_notes",
]


# Header display names — internal column key → user-friendly label
COLUMN_HEADER_DISPLAY = {
    "slot": "Trade #",
}


def _fmt_expiry(val: object) -> str:
    """Format ISO date as '9th Jul 26' style."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    try:
        from datetime import date as _date
        s = str(val)[:10]
        d = _date.fromisoformat(s)
        day = d.day
        if 11 <= day <= 13:
            sfx = "th"
        else:
            sfx = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
        return f"{day}{sfx} {d.strftime('%b')} {d.strftime('%y')}"
    except Exception:
        return str(val)


def _fmt_exit_time(val: object) -> str:
    """Format ISO timestamp as 'DD-Mon HH:MM' IST (e.g. '05-Jun 14:25')."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    try:
        ts = pd.to_datetime(str(val))
        if ts.tzinfo is None:
            ts = ts.tz_localize(IST)
        else:
            ts = ts.tz_convert(IST)
        return ts.strftime("%d-%b %H:%M")
    except Exception:
        return str(val)


def _fmt_is_expiry_day(val: object) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, bool):
        return "Expiry Day" if val else "Normal Day"
    s = str(val).strip().lower()
    if s in ("true", "1", "yes"):
        return "Expiry Day"
    if s in ("false", "0", "no"):
        return "Normal Day"
    return str(val)


def _truncate(val: object, n: int = 60) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val)
    return s if len(s) <= n else s[: n - 1] + "…"


def build_paper_trades_sheet(
    ws: Worksheet,
    paper_trades_df: pd.DataFrame,
) -> int:
    """Fill the ``Paper Trades`` sheet from a merged-overrides frame.

    The frame must already include ``manual_*`` columns (filled or
    None) — call ``src.paper.persistence.merge_overrides`` before
    handing it here.
    """
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1, value="Paper Trades")
    title.font = PAPER_TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    sub = ws.cell(row=2, column=1, value=PAPER_SHEET_BANNER)
    sub.font = PAPER_SUB_FONT

    # Quick help box — explains terms NOT shown in the sheet but still
    # in the JSONL. Update PAPER_TRADES_HEADER_ROW below if rows change.
    ws.merge_cells("A3:H3")
    help_title = ws.cell(row=3, column=1, value="📖 Quick help")
    help_title.font = Font(name="Calibri", bold=True, color="2E7D32", size=11)
    help_lines = [
        "• Trade #  →  trade number for the day (cap is 3 per config).",
        "• result_chip  →  🟢 TP2 / TP1   🟡 TP1→BE   🔴 SL / Hard   ⏹ SqOff (3:00 PM)   · N/A (no data yet).",
        "• intrabar_ambiguous (in JSONL, not shown)  →  same 5-min candle hit BOTH SL and TP.",
        "    e.g. entry ₹100, SL ₹90, TP ₹110; candle low ₹85, high ₹115 → can't tell which side hit first from OHLC.",
        "• fidelity (in JSONL, not shown)  →  'ohlc' is normal (full data). 'close_only' = legacy log, MFE/MAE less precise.",
    ]
    for i, line in enumerate(help_lines):
        r = 4 + i
        ws.merge_cells(f"A{r}:H{r}")
        cell = ws.cell(row=r, column=1, value=line)
        cell.font = Font(name="Calibri", italic=True, color="595959", size=10)

    # Headers live at row 10 (1 title + 1 banner + 1 help-title + 5 help + 1 blank).
    header_row = PAPER_TRADES_HEADER_ROW

    if paper_trades_df is None or paper_trades_df.empty:
        _set_paper_headers(ws, ["(no paper trades yet — run the bot)"], row=header_row)
        return 0

    df = paper_trades_df.copy()
    # Only show trades we actually took — SKIPPED rows are not paper trades.
    if "decision" in df.columns:
        df = df[df["decision"].str.upper() == "TAKEN"].copy()
    if df.empty:
        _set_paper_headers(ws, ["(no TAKEN trades yet)"], row=header_row)
        return 0
    for col in PAPER_TRADE_COLUMNS:
        if col not in df.columns:
            df[col] = None

    # Compute the result chip column at write-time so the dashboard
    # never disagrees with the underlying outcome / decision.
    df["result_chip"] = df.apply(
        lambda r: _outcome_chip(r.get("outcome"), r.get("decision")),
        axis=1,
    )

    df = df[PAPER_TRADE_COLUMNS].copy()
    if "expiry" in df.columns:
        df["expiry"] = df["expiry"].apply(_fmt_expiry)
    if "exit_time" in df.columns:
        df["exit_time"] = df["exit_time"].apply(_fmt_exit_time)
    if "is_expiry_day" in df.columns:
        df["is_expiry_day"] = df["is_expiry_day"].apply(_fmt_is_expiry_day)
    if "bot_remark" in df.columns:
        df["bot_remark"] = df["bot_remark"].apply(_truncate)
    headers = [COLUMN_HEADER_DISPLAY.get(c, c) for c in df.columns]
    _set_paper_headers(ws, headers, row=header_row)

    written = 0
    for offset, (_, row) in enumerate(df.iterrows()):
        excel_row = header_row + 1 + offset
        for col_idx, col in enumerate(df.columns, start=1):
            v = row[col]
            if isinstance(v, list):
                v = ",".join(str(x) for x in v) if v else None
            if pd.isna(v):
                v = None
            cell = ws.cell(row=excel_row, column=col_idx, value=v)
            cell.font = PAPER_BODY_FONT
        written += 1

    _autofit(ws, df, header_row=header_row)

    # ---------------- per-row fills by outcome / decision ----------------
    col_index = {name: idx + 1 for idx, name in enumerate(df.columns)}
    n_cols = len(df.columns)

    for offset, (_, row) in enumerate(df.iterrows()):
        excel_row = header_row + 1 + offset
        decision = str(row.get("decision") or "").upper()
        outcome = str(row.get("outcome") or "").upper()
        fill = None
        strikethrough = False
        if decision == "SKIPPED":
            fill = PAPER_SKIPPED_FILL
            strikethrough = True
        else:
            fill = PAPER_OUTCOME_FILLS.get(outcome)
        if fill is not None:
            for c in range(1, n_cols + 1):
                ws.cell(row=excel_row, column=c).fill = fill
        if strikethrough:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=excel_row, column=c)
                cell.font = Font(name="Calibri", size=11, strike=True)

    # ---------------- data bars on realized_R + paper_pnl ----------------
    if written > 0:
        for col_name, positive_color, negative_color in [
            ("realized_R", "63B881", "EF6C6C"),
            ("paper_pnl", "63B881", "EF6C6C"),
        ]:
            if col_name not in col_index:
                continue
            letter = get_column_letter(col_index[col_name])
            range_ref = f"{letter}{header_row + 1}:{letter}{header_row + written}"
            ws.conditional_formatting.add(
                range_ref,
                DataBarRule(
                    start_type="num", start_value=-3,
                    end_type="num", end_value=3,
                    color=positive_color,
                ),
            )
            ws.conditional_formatting.add(
                range_ref,
                CellIsRule(
                    operator="lessThan", formula=["0"],
                    fill=PatternFill("solid", fgColor=negative_color),
                ),
            )

    # ---------------- data-validation dropdowns on manual columns -------
    if written > 0:
        dec_letter = get_column_letter(col_index["manual_decision"])
        out_letter = get_column_letter(col_index["manual_outcome"])
        dec_range = f"{dec_letter}{header_row + 1}:{dec_letter}{header_row + written}"
        out_range = f"{out_letter}{header_row + 1}:{out_letter}{header_row + written}"

        dec_dv = DataValidation(
            type="list",
            formula1=f'"{MANUAL_DECISION_CHOICES}"',
            allow_blank=True,
            showErrorMessage=False,
        )
        dec_dv.add(dec_range)
        ws.add_data_validation(dec_dv)

        out_dv = DataValidation(
            type="list",
            formula1=f'"{MANUAL_OUTCOME_CHOICES}"',
            allow_blank=True,
            showErrorMessage=False,
        )
        out_dv.add(out_range)
        ws.add_data_validation(out_dv)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    return written


# ---------------------------------------------------------------------------
# Paper Dashboard (KPIs via Excel formulas)
# ---------------------------------------------------------------------------


def build_paper_dashboard_sheet(
    ws: Worksheet,
    paper_trades_df: pd.DataFrame,
    paper_trades_sheet_name: str = "Paper Trades",
    collapse_summary: str | None = None,
) -> None:
    """KPI cards built with Excel formulas over the Paper Trades sheet.

    The dashboard reads manual-first then auto: every KPI references
    a helper column whose value is ``manual_outcome`` if set else
    ``outcome``. That helper column lives on Paper Trades and is
    populated at write-time so the formulas don't need IFERROR.
    """
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1, value="Paper Dashboard")
    title.font = PAPER_TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    sub = ws.cell(row=2, column=1, value=PAPER_SHEET_BANNER)
    sub.font = PAPER_SUB_FONT

    if paper_trades_df is None or paper_trades_df.empty:
        ws.cell(row=4, column=1, value="(no paper trades yet)").font = PAPER_SUB_FONT
        return

    # Range references — Paper Trades data block starts after the
    # title + banner + help box. Single source of truth lives above.
    n = len(paper_trades_df)
    data_start = PAPER_TRADES_DATA_START
    data_end = data_start + n - 1
    sheet_ref = f"'{paper_trades_sheet_name}'!"

    # Map columns to letters. Must match the column order in
    # ``build_paper_trades_sheet`` exactly. The Paper Trades sheet now
    # contains ONLY TAKEN rows, so formulas no longer filter on decision.
    column_letter = {
        col: get_column_letter(i + 1)
        for i, col in enumerate(PAPER_TRADE_COLUMNS)
    }
    date_col = column_letter["date"]
    symbol_col = column_letter["symbol"]
    outcome_col = column_letter["outcome"]
    realized_R_col = column_letter["realized_R"]
    paper_pnl_col = column_letter["paper_pnl"]

    def rng(col_letter: str) -> str:
        return f"{sheet_ref}{col_letter}{data_start}:{col_letter}{data_end}"

    out_rng = rng(outcome_col)
    pnl_rng = rng(paper_pnl_col)
    sym_rng = rng(symbol_col)
    date_rng = rng(date_col)
    R_rng = rng(realized_R_col)

    # Resolved = every TAKEN trade whose outcome is not pending (NO_DATA).
    total_taken_f = f"=COUNTA({date_rng})"
    resolved_f = f"=COUNTA({out_rng})-COUNTIF({out_rng},\"NO_DATA\")"
    wins_f = f"=COUNTIF({out_rng},\"TP2_HIT\")+COUNTIF({out_rng},\"TP1_HIT\")"
    sl_count_f = f"=COUNTIF({out_rng},\"SL_HIT\")"
    no_data_count_f = f"=COUNTIF({out_rng},\"NO_DATA\")"
    headline_pnl_f = f"=SUM({pnl_rng})"
    avg_R_f = (
        f"=IFERROR(AVERAGEIFS({R_rng},{out_rng},\"<>NO_DATA\"),0)"
    )
    win_rate_f = (
        f"=IFERROR(({wins_f.lstrip('=')})/({resolved_f.lstrip('=')}),0)"
    )
    sl_rate_f = (
        f"=IFERROR(({sl_count_f.lstrip('=')})/({resolved_f.lstrip('=')}),0)"
    )
    best_pnl_f = f"=IFERROR(MAX({pnl_rng}),0)"
    worst_pnl_f = f"=IFERROR(MIN({pnl_rng}),0)"
    best_detail_f = (
        f"=IFERROR(INDEX({sym_rng},MATCH(MAX({pnl_rng}),{pnl_rng},0))"
        f"&\" — \"&INDEX({date_rng},MATCH(MAX({pnl_rng}),{pnl_rng},0)),\"\")"
    )
    worst_detail_f = (
        f"=IFERROR(INDEX({sym_rng},MATCH(MIN({pnl_rng}),{pnl_rng},0))"
        f"&\" — \"&INDEX({date_rng},MATCH(MIN({pnl_rng}),{pnl_rng},0)),\"\")"
    )

    kpi_rows = [
        ("Total Trades Taken",   total_taken_f,   None),
        ("Win Rate",             win_rate_f,      None),
        ("Total Paper P&L (₹)",  headline_pnl_f,  None),
        ("Avg R per Trade",      avg_R_f,         None),
        ("Best Trade (₹)",       best_pnl_f,      best_detail_f),
        ("Worst Trade (₹)",      worst_pnl_f,     worst_detail_f),
        ("SL Hit Rate",          sl_rate_f,       None),
        ("Pending (NO_DATA)",    no_data_count_f, None),
    ]

    # Headline P&L (TAKEN only) — large + bold, deep green.
    ws.cell(row=4, column=1, value="HEADLINE PAPER P&L (TAKEN ONLY)").font = Font(
        bold=True, color="FFFFFF", size=11
    )
    ws.cell(row=4, column=1).fill = PatternFill("solid", fgColor="2E7D32")
    ws.merge_cells("A4:C4")

    big = ws.cell(row=5, column=1, value=headline_pnl_f)
    big.font = Font(size=22, bold=True, color="2E7D32")
    big.alignment = Alignment(horizontal="left", vertical="center")
    big.number_format = '₹#,##0.00;[Red]-₹#,##0.00'
    ws.merge_cells("A5:C5")
    ws.row_dimensions[5].height = 36

    # KPI table.
    row = 7
    ws.cell(row=row, column=1, value="KPI").font = PAPER_HEADER_FONT
    ws.cell(row=row, column=1).fill = PAPER_HEADER_FILL
    ws.cell(row=row, column=2, value="Value").font = PAPER_HEADER_FONT
    ws.cell(row=row, column=2).fill = PAPER_HEADER_FILL
    ws.cell(row=row, column=3, value="Detail").font = PAPER_HEADER_FONT
    ws.cell(row=row, column=3).fill = PAPER_HEADER_FILL
    for i, (label, formula, detail) in enumerate(kpi_rows, start=1):
        ws.cell(row=row + i, column=1, value=label).font = PAPER_BODY_FONT
        cell = ws.cell(row=row + i, column=2, value=formula)
        cell.font = PAPER_BODY_FONT
        if "rate" in label.lower():
            cell.number_format = "0.0%"
        elif "₹" in label:
            cell.number_format = '₹#,##0'
        elif label == "Avg R per Trade":
            cell.number_format = "0.00"
        if detail is not None:
            d = ws.cell(row=row + i, column=3, value=detail)
            d.font = PAPER_BODY_FONT

    # Outcome distribution table + collapse ratio.
    table_start = row + len(kpi_rows) + 3
    ws.cell(row=table_start, column=1, value="Outcome distribution").font = Font(
        bold=True, color="2E7D32", size=12
    )
    ws.cell(row=table_start + 1, column=1, value="Outcome").font = PAPER_HEADER_FONT
    ws.cell(row=table_start + 1, column=1).fill = PAPER_HEADER_FILL
    ws.cell(row=table_start + 1, column=2, value="Count").font = PAPER_HEADER_FONT
    ws.cell(row=table_start + 1, column=2).fill = PAPER_HEADER_FILL
    for j, outcome_label in enumerate(
        ["TP2_HIT", "TP1_HIT", "TP1_BE", "SL_HIT", "HARD_EXIT", "OPEN_SQOFF", "NO_DATA"],
        start=2,
    ):
        ws.cell(row=table_start + j, column=1, value=outcome_label).font = PAPER_BODY_FONT
        fill = PAPER_OUTCOME_FILLS.get(outcome_label)
        if fill:
            ws.cell(row=table_start + j, column=1).fill = fill
        ws.cell(
            row=table_start + j, column=2,
            value=f'=COUNTIF({rng(outcome_col)},"{outcome_label}")',
        ).font = PAPER_BODY_FONT

    # Collapse-ratio block (so re-fire inflation is visible).
    cr_row = table_start + 11
    ws.cell(row=cr_row, column=1, value="Episode collapse ratio").font = Font(
        bold=True, color="2E7D32", size=12
    )
    ws.cell(row=cr_row + 1, column=1, value=(collapse_summary or "(not recorded)"))
    ws.cell(row=cr_row + 1, column=1).font = PAPER_BODY_FONT

    # Generated-at footer.
    foot_row = cr_row + 3
    ws.cell(
        row=foot_row, column=1,
        value=f"Generated at {datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S')} IST",
    ).font = PAPER_SUB_FONT

    # Column widths
    for col_letter in ("A", "B", "C"):
        ws.column_dimensions[col_letter].width = 32


# ---------------------------------------------------------------------------
# Echoes sheet (hidden by default)
# ---------------------------------------------------------------------------


def build_echoes_sheet(
    ws: Worksheet,
    annotated_alerts: pd.DataFrame,
) -> int:
    """Echoes — every non-representative re-fire.

    Hidden by default so the main Paper Trades sheet stays
    first-alert-only. Kept for diagnostics: re-fire inflation, per-
    strike clustering, late chasers, etc.
    """
    ws.sheet_state = "hidden"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1, value="Echoes (diagnostic)")
    title.font = PAPER_TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:H2")
    sub = ws.cell(
        row=2, column=1,
        value=(
            "Non-representative alerts — every re-fire that was collapsed "
            "into an episode. NOT counted in TAKEN P&L."
        ),
    )
    sub.font = PAPER_SUB_FONT

    header_row = 4
    if annotated_alerts is None or annotated_alerts.empty:
        _set_paper_headers(ws, ["(no alerts loaded)"], row=header_row)
        return 0

    echoes = annotated_alerts[annotated_alerts["paper_role"] == "echo"].copy()
    if echoes.empty:
        _set_paper_headers(ws, ["(no echoes — no re-fires)"], row=header_row)
        return 0

    show_cols = [
        c for c in (
            "alert_id", "episode_id", "candle_ts", "date", "symbol",
            "strike", "relation", "option_type", "entry", "sl",
            "tp1", "tp2", "bot_remark", "bot_tags", "fidelity_note",
        ) if c in echoes.columns
    ]
    out = echoes[show_cols].copy()
    if "candle_ts" in out.columns:
        out["candle_ts"] = out["candle_ts"].astype(str)
    out = out.sort_values(by=["episode_id", "candle_ts"]).reset_index(drop=True)

    _set_paper_headers(ws, out.columns, row=header_row)
    for offset, (_, row) in enumerate(out.iterrows()):
        excel_row = header_row + 1 + offset
        for col_idx, col in enumerate(out.columns, start=1):
            v = row[col]
            if pd.isna(v):
                v = None
            ws.cell(row=excel_row, column=col_idx, value=v).font = PAPER_BODY_FONT
    _autofit(ws, out, header_row=header_row)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    return len(out)
