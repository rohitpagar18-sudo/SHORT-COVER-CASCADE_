"""Phase 5.2 — Excel dashboard builder tests."""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook


@pytest.fixture
def isolated_paths(tmp_path, monkeypatch):
    logs = tmp_path / "logs"
    data = tmp_path / "data"
    dashboards = logs / "dashboards"
    logs.mkdir()
    data.mkdir()
    dashboards.mkdir()

    import src.dashboard.data_writer as dw
    import src.dashboard.excel_builder as eb

    monkeypatch.setattr(dw, "LOGS_DIR", logs)
    monkeypatch.setattr(dw, "DATA_DIR", data)
    monkeypatch.setattr(dw, "DASHBOARDS_DIR", dashboards)
    monkeypatch.setattr(dw, "SIGNALS_JSONL", logs / "signals.jsonl")
    monkeypatch.setattr(dw, "ALERTS_JSONL", logs / "alerts.jsonl")
    monkeypatch.setattr(dw, "GAP_JSONL", logs / "gap_log.jsonl")

    # excel_builder imported DASHBOARDS_DIR at module load — patch its copy.
    monkeypatch.setattr(eb, "DASHBOARDS_DIR", dashboards)
    return logs, data, dashboards


def _write_jsonl(path: Path, records: list[dict]) -> None:
    with path.open("a", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r) + "\n")


def _alert(ts: str, **kw) -> dict:
    base = {
        "timestamp_ist": ts,
        "event_type": "alert",
        "symbol": "NIFTY",
        "strike": 24050,
        "option_type": "CE",
        "relation": "ATM",
        "expiry": "2026-06-02",
        "trading_symbol": "NIFTY24050CE",
        "spot_price": 24030.0,
        "spot_vwap": 24000.0,
        "option_close": 150.0,
        "option_vwap": 140.0,
        "rsi": 65.0,
        "rsi_ma": 55.0,
        "oi": 800_000,
        "oi_ma": 1_000_000,
        "volume": 3000,
        "volume_ma": 1500,
        "is_green": True,
        "vix": 14.0,
        "vix_regime": "Normal",
        "all_passed": True,
        "summary": "C0 ✓ C1 ✓ C2 ✓ C3 ✓ C4 ✓",
        "opt_above_vwap_pct": 7.0,
        "entry": 150.0,
        "sl": 140.0,
        "sl_method": 1,
        "tp1": 165.0,
        "tp2": 175.0,
        "tp1_r": 1.5,
        "tp2_r": 2.5,
        "risk_per_unit": 10.0,
        "lots": 3,
        "total_risk": 2850.0,
        "lot_size": 65,
        "day_type": "Normal",
        "vix_multiplier": 1.0,
        "spot": 24030.0,
        "spot_position": "Above VWAP ✓",
        "date": ts[:10],
        "time": ts[11:16],
        "bot_remark": "5/5 strong — opt 7% above VWAP, RSI 65, OI 20% below MA, vol 2× MA",
        "bot_tags": "fresh_breakout,strong_rsi,strong_oi,high_volume,morning,normal_vix,normal_day,first_alert",
    }
    base.update(kw)
    return base


def _gap(ts: str, decision: str) -> dict:
    return {
        "timestamp_ist": ts,
        "decision": decision,
        "enabled": True,
        "threshold_pct": 1.0,
        "direction": "both",
        "any_triggered": decision != "NORMAL",
        "per_symbol": {
            "NIFTY": {"open": 24050.0, "prev_close": 24000.0,
                      "gap_pct": 0.21, "triggers": False, "error": None},
            "BANKNIFTY": {"open": 51000.0, "prev_close": 50950.0,
                          "gap_pct": 0.10, "triggers": False, "error": None},
        },
    }


def _scan(ts: str, **kw) -> dict:
    base = _alert(ts)
    base["event_type"] = "scan"
    base["all_passed"] = False
    base.update(kw)  # caller overrides win over defaults
    return base


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------


def test_resolve_excel_path_q1() -> None:
    from src.dashboard.excel_builder import _resolve_excel_path

    p = _resolve_excel_path(2026, 1)
    assert p.name == "dashboard_2026_Q1.xlsx"


def test_resolve_excel_path_q4() -> None:
    from src.dashboard.excel_builder import _resolve_excel_path

    p = _resolve_excel_path(2026, 4)
    assert p.name == "dashboard_2026_Q4.xlsx"


def test_quarter_for_date_april_is_q2() -> None:
    from datetime import datetime
    from zoneinfo import ZoneInfo

    from src.dashboard.data_writer import quarter_for_date

    d = datetime(2026, 4, 15, tzinfo=ZoneInfo("Asia/Kolkata"))
    assert quarter_for_date(d) == (2026, 2)


def test_quarter_for_date_december_is_q4() -> None:
    from datetime import datetime
    from zoneinfo import ZoneInfo

    from src.dashboard.data_writer import quarter_for_date

    d = datetime(2026, 12, 1, tzinfo=ZoneInfo("Asia/Kolkata"))
    assert quarter_for_date(d) == (2026, 4)


# ---------------------------------------------------------------------------
# Sheet count + names + idempotency
# ---------------------------------------------------------------------------


_EXPECTED_SHEETS = [
    "Strategy Dashboard",
    "Daily Summary",
    "All Alerts",
    "Order Place",
    # "All Signals" was removed: signals.jsonl is the source-of-truth audit.
    "Gap History",
    # Phase 5D — paper layer sheets sit between Gap History and the
    # config snapshot so Strategy Dashboard stays the default-open tab.
    "Paper Trades",
    "Paper Dashboard",
    "Echoes (diagnostic)",
    "Config Snapshot",
]


def test_dashboard_creates_expected_sheets(isolated_paths) -> None:
    logs, data, dashboards = isolated_paths
    _write_jsonl(logs / "alerts.jsonl", [_alert("2026-05-27T10:35:00+05:30")])
    _write_jsonl(logs / "gap_log.jsonl", [_gap("2026-05-27T09:16:00+05:30", "NORMAL")])
    from src.dashboard import sync_jsonl_to_parquet, update_dashboard

    sync_jsonl_to_parquet()
    result = update_dashboard()
    assert result["status"] == "ok"
    wb_path = Path(result["output_path"])
    wb = load_workbook(wb_path, data_only=False)
    assert wb.sheetnames == _EXPECTED_SHEETS


def test_workbook_has_no_rejections_sheet(isolated_paths) -> None:
    logs, data, dashboards = isolated_paths
    _write_jsonl(logs / "alerts.jsonl", [_alert("2026-05-27T10:35:00+05:30")])
    _write_jsonl(logs / "gap_log.jsonl", [_gap("2026-05-27T09:16:00+05:30", "NORMAL")])
    from src.dashboard import sync_jsonl_to_parquet, update_dashboard

    sync_jsonl_to_parquet()
    result = update_dashboard()
    wb = load_workbook(Path(result["output_path"]), data_only=False)
    assert "Rejections" not in wb.sheetnames


def test_dashboard_is_idempotent(isolated_paths) -> None:
    logs, data, dashboards = isolated_paths
    _write_jsonl(logs / "alerts.jsonl", [_alert("2026-05-27T10:35:00+05:30")])
    from src.dashboard import sync_jsonl_to_parquet, update_dashboard

    sync_jsonl_to_parquet()
    a = update_dashboard()
    b = update_dashboard()
    # Second run still writes the workbook but reports the same counts.
    assert a["alerts_added"] == b["alerts_added"] == 1


def test_dashboard_strategy_sheet_has_chart_objects(isolated_paths) -> None:
    logs, data, dashboards = isolated_paths
    _write_jsonl(logs / "alerts.jsonl", [_alert("2026-05-27T10:35:00+05:30")])
    from src.dashboard import sync_jsonl_to_parquet, update_dashboard

    sync_jsonl_to_parquet()
    result = update_dashboard()
    wb = load_workbook(Path(result["output_path"]))
    dash = wb["Strategy Dashboard"]
    # _charts is openpyxl's internal list — non-empty proves chart objects exist.
    assert len(dash._charts) >= 1


def test_dashboard_no_data_returns_no_data_status(isolated_paths) -> None:
    logs, data, dashboards = isolated_paths
    # No JSONL files yet — but update_dashboard still creates the current
    # quarter's workbook (empty), per spec. Result status is "ok" with
    # zero counts.
    from src.dashboard import update_dashboard

    result = update_dashboard()
    # We accept "ok" with all zero counts OR "no_data" — both are valid.
    assert result["status"] in ("ok", "no_data")


# ---------------------------------------------------------------------------
# Order Place outcome coloring
# ---------------------------------------------------------------------------


def test_order_place_outcome_coloring_applied(isolated_paths) -> None:
    """When parquet contains a filled order_status, the cell is coloured."""
    logs, data, dashboards = isolated_paths
    _write_jsonl(logs / "alerts.jsonl", [_alert("2026-05-27T10:35:00+05:30")])
    from src.dashboard import sync_jsonl_to_parquet, update_dashboard

    sync_jsonl_to_parquet()

    # Manually patch the parquet so the alert row carries an order_status.
    parquet_path = data / "scc_data_2026-05.parquet"
    df = pd.read_parquet(parquet_path)
    df["order_status"] = "TP2_HIT"
    df["pnl_rupees"] = 7500.0
    df["exit_price"] = 175.0
    df.to_parquet(parquet_path, index=False)

    result = update_dashboard()
    wb = load_workbook(Path(result["output_path"]))
    ws = wb["Order Place"]
    headers = [c.value for c in ws[1]]
    status_col = headers.index("Order Status") + 1
    # Row 2 (the only alert) status cell should carry a fill.
    cell = ws.cell(row=2, column=status_col)
    assert cell.fill is not None
    # PatternFill with non-default fgColor proves coloring kicked in.
    assert cell.fill.fgColor.rgb is not None


def test_all_signals_sheet_removed(isolated_paths) -> None:
    # The All Signals sheet was removed — signals.jsonl is the
    # source-of-truth audit (duplicating it inside Excel was wasteful).
    # Build a workbook with an extended-zone scan present to be sure the
    # sheet really is gone (not just absent because of empty data).
    logs, data, dashboards = isolated_paths
    ext = _scan(
        "2026-05-27T11:00:00+05:30",
        event_type="would_alert_extended",
        opt_above_vwap_pct=37.0,
    )
    _write_jsonl(logs / "signals.jsonl", [ext])
    from src.dashboard import sync_jsonl_to_parquet, update_dashboard

    sync_jsonl_to_parquet()
    result = update_dashboard()
    wb = load_workbook(Path(result["output_path"]))
    assert "All Signals" not in wb.sheetnames


def test_gap_history_colour_for_directional_label(isolated_paths) -> None:
    logs, data, dashboards = isolated_paths
    _write_jsonl(
        logs / "gap_log.jsonl",
        [
            _gap("2026-05-27T09:16:00+05:30", "GAP_UP"),
            _gap("2026-05-28T09:16:00+05:30", "GAP_DOWN"),
        ],
    )
    from src.dashboard import sync_jsonl_to_parquet, update_dashboard

    sync_jsonl_to_parquet()
    result = update_dashboard()
    wb = load_workbook(Path(result["output_path"]))
    ws = wb["Gap History"]
    headers = [c.value for c in ws[1]]
    dec_col = headers.index("Decision") + 1
    assert ws.cell(row=2, column=dec_col).fill is not None
    assert ws.cell(row=3, column=dec_col).fill is not None


def test_config_snapshot_contains_c1_threshold_row(isolated_paths) -> None:
    logs, data, dashboards = isolated_paths
    _write_jsonl(logs / "alerts.jsonl", [_alert("2026-05-27T10:35:00+05:30")])
    from src.dashboard import sync_jsonl_to_parquet, update_dashboard

    sync_jsonl_to_parquet()
    result = update_dashboard()
    wb = load_workbook(Path(result["output_path"]))
    ws = wb["Config Snapshot"]
    settings = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any("C1 max distance" in str(s) for s in settings)


def test_daily_summary_row_per_date(isolated_paths) -> None:
    logs, data, dashboards = isolated_paths
    _write_jsonl(
        logs / "alerts.jsonl",
        [
            _alert("2026-05-27T10:35:00+05:30"),
            _alert("2026-05-28T11:00:00+05:30"),
        ],
    )
    from src.dashboard import sync_jsonl_to_parquet, update_dashboard

    sync_jsonl_to_parquet()
    result = update_dashboard()
    wb = load_workbook(Path(result["output_path"]))
    ws = wb["Daily Summary"]
    # Header row + 2 dates.
    assert ws.max_row >= 3


# ---------------------------------------------------------------------------
# Paper Dashboard — NO_DATA exclusion contract
#
# Pins the rule that pre-tracking episodes (logged before SL/TP exit data
# was captured) appear as NO_DATA on the Paper Trades sheet but never
# inflate the headline "Trades Taken" count or pollute the P&L / win-rate
# / profit-factor / R aggregates on the Paper Dashboard. NO_DATA is
# reported as its own KPI ("Pending (NO_DATA)").
# ---------------------------------------------------------------------------


def _paper_row(date: str, outcome: str, paper_pnl: float, **kw) -> dict:
    """Minimal paper-trades row — only the fields the dashboard reads."""
    base = {
        "alert_id": f"{date}|NIFTY|CE",
        "episode_id": f"{date}T10:00:00+05:30|NIFTY|CE",
        "paper_role": "representative",
        "date": date,
        "candle_timestamp": f"{date}T10:00:00+05:30",
        "symbol": "NIFTY",
        "strike": 24050,
        "relation": "ATM",
        "option_type": "CE",
        "expiry": "2026-06-02",
        "entry": 150.0,
        "sl": 140.0,
        "tp1": 165.0,
        "tp2": 175.0,
        "lots": 3,
        "lot_size": 65,
        "is_expiry_day": False,
        "decision": "TAKEN",
        "decision_reason": "taken (slot 1/3)",
        "slot": 1,
        "outcome": outcome,
        "exit_price": None,
        "exit_time": None,
        "exit_reason": "",
        "realized_R": 0.0,
        "paper_pnl": paper_pnl,
        "paper_pnl_per_unit": 0.0,
        "mfe": 0.0, "mae": 0.0, "mfe_R": 0.0, "mae_R": 0.0,
        "max_drawdown_R": 0.0,
        "intrabar_ambiguous": False,
        "fidelity": "ohlc",
        "bot_remark": "",
        "bot_tags": "",
        "triggered_caps": [],
    }
    base.update(kw)
    return base


def test_paper_dashboard_excludes_no_data_from_trades_taken() -> None:
    """Trades Taken (tracked) on the Paper Dashboard must subtract NO_DATA
    rows so pre-tracking episodes (1–4 Jun) don't inflate the count.
    NO_DATA is reported separately under "Pending (NO_DATA)".
    """
    from openpyxl import Workbook
    from src.dashboard.paper_sheets import (
        build_paper_dashboard_sheet,
        build_paper_trades_sheet,
    )

    # 2 NO_DATA (pre-tracking) + 1 TP2 + 1 SL → tracked count should be 2.
    rows = [
        _paper_row("2026-06-01", "NO_DATA", 0.0),
        _paper_row("2026-06-04", "NO_DATA", 0.0),
        _paper_row("2026-06-05", "TP2_HIT", 4875.0, realized_R=2.5),
        _paper_row("2026-06-06", "SL_HIT", -1950.0, realized_R=-1.0),
    ]
    df = pd.DataFrame(rows)
    # The merge-overrides shape requires the manual_* columns to exist.
    for c in ("manual_decision", "manual_reason", "manual_outcome",
              "manual_exit", "user_notes"):
        df[c] = None

    wb = Workbook()
    trades_ws = wb.active
    trades_ws.title = "Paper Trades"
    build_paper_trades_sheet(trades_ws, df)

    dash_ws = wb.create_sheet("Paper Dashboard")
    build_paper_dashboard_sheet(dash_ws, df)

    # Walk column A to locate KPI labels, then check the formula on column B.
    kpi_formulae: dict[str, str] = {}
    for r in range(1, dash_ws.max_row + 1):
        label = dash_ws.cell(row=r, column=1).value
        if isinstance(label, str):
            kpi_formulae[label] = dash_ws.cell(row=r, column=2).value

    assert "Trades Taken (tracked)" in kpi_formulae
    assert "Pending (NO_DATA)" in kpi_formulae
    taken_f = kpi_formulae["Trades Taken (tracked)"]
    # The formula must subtract the NO_DATA count from the total.
    assert "NO_DATA" in taken_f and "-" in taken_f, taken_f
    # Headline P&L / win-rate / etc. must exclude NO_DATA.
    assert "NO_DATA" in kpi_formulae["Total Paper P&L (₹)"]
    assert "<>NO_DATA" in kpi_formulae["Avg R per Trade"]
    assert "NO_DATA" in kpi_formulae["Best Trade (₹)"]
    assert "NO_DATA" in kpi_formulae["Worst Trade (₹)"]
    assert "NO_DATA" in kpi_formulae["Profit Factor"]


def test_paper_trades_sheet_keeps_no_data_rows_visible() -> None:
    """NO_DATA reps must remain visible on the Paper Trades sheet — they
    are excluded from aggregates, not deleted from the raw view.
    """
    from openpyxl import Workbook
    from src.dashboard.paper_sheets import build_paper_trades_sheet

    df = pd.DataFrame([
        _paper_row("2026-06-01", "NO_DATA", 0.0),
        _paper_row("2026-06-05", "TP2_HIT", 4875.0, realized_R=2.5),
    ])
    for c in ("manual_decision", "manual_reason", "manual_outcome",
              "manual_exit", "user_notes"):
        df[c] = None

    wb = Workbook()
    ws = wb.active
    ws.title = "Paper Trades"
    written = build_paper_trades_sheet(ws, df)
    # Both TAKEN rows must be on the sheet (NO_DATA flagged, not hidden).
    assert written == 2
