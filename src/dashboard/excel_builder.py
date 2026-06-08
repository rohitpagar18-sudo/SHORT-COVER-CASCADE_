"""Phase 5.2 — Quarterly Excel dashboard builder.

Reads the unified monthly Parquet files in ``data/`` and produces a
human-facing workbook at ``logs/dashboards/dashboard_YYYY_QN.xlsx``.

Seven sheets in order:

1. Strategy Dashboard — KPI tiles + 4 charts (visual at-a-glance view)
2. Daily Summary       — one row per trading day with the directional
                         gap label and aggregate counts
3. All Alerts          — every 5/5 alert; includes opt_above_vwap_pct,
                         bot_remark, bot_tags
4. Order Place         — automatic columns + manual columns the user
                         fills in (order status, exit price, P&L,
                         user notes); coloured by outcome
5. All Signals         — full audit including ``would_alert_extended``
                         rows highlighted in light orange
6. Gap History         — directional labels with colour swatches
7. Config Snapshot     — current config values for this quarter

Idempotent: re-running ``update_dashboard`` will rebuild the workbook
from scratch using the latest Parquet state.

User-facing extras (Phase 5.2 stretch):
- Headers are coloured; KPI tiles have a coloured top stripe.
- The "Insight:" column on All Alerts is left-aligned and wrapped.
- Order Place sheet has an in-sheet outcome-category legend.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

import pandas as pd
from loguru import logger

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.dashboard.data_writer import (
    DASHBOARDS_DIR,
    load_parquet_for_quarter,
    quarter_for_date,
)

IST = ZoneInfo("Asia/Kolkata")

# ---------------------------------------------------------------------------
# Style palette
# ---------------------------------------------------------------------------

# Header background + font
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")  # deep navy
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(name="Calibri", bold=True, color="1F4E78", size=16)
_SUBTITLE_FONT = Font(name="Calibri", italic=True, color="595959", size=10)
_BODY_FONT = Font(name="Calibri", size=11)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)

# Outcome colour palette
OUTCOME_FILLS = {
    "TP2_HIT": PatternFill("solid", fgColor="6FBF73"),      # strong green
    "TP1_HIT": PatternFill("solid", fgColor="C8E6C9"),      # light green
    "SL_HIT": PatternFill("solid", fgColor="EF9A9A"),       # red
    "WOULD_SKIP": PatternFill("solid", fgColor="E0E0E0"),   # grey
    "PARTIAL": PatternFill("solid", fgColor="FFE082"),      # yellow
}

# Gap decision colours
GAP_FILLS = {
    "GAP_UP": PatternFill("solid", fgColor="FFCDD2"),            # light red
    "GAP_DOWN": PatternFill("solid", fgColor="BBDEFB"),          # light blue
    "GAP_UP_DISABLED": PatternFill("solid", fgColor="FFE0B2"),   # orange
    "GAP_DOWN_DISABLED": PatternFill("solid", fgColor="BBDEFB"), # light blue
    # Back-compat labels still found in legacy gap_log rows:
    "GAP_DAY": PatternFill("solid", fgColor="FFCDD2"),
    "GAP_DETECTED_BUT_DISABLED": PatternFill("solid", fgColor="FFE0B2"),
}

# Signal row colour by Relation. Depths share a family colour:
# ITM = grey, ATM = yellow, OTM = purple. Deeper levels darken.
_ITM_FILL = PatternFill("solid", fgColor="F4F4F4")
_ITM2_FILL = PatternFill("solid", fgColor="E0E0E0")
_ITM3_FILL = PatternFill("solid", fgColor="BDBDBD")
_ATM_FILL = PatternFill("solid", fgColor="FFF9C4")
_OTM_FILL = PatternFill("solid", fgColor="EDE7F6")
_OTM2_FILL = PatternFill("solid", fgColor="D1C4E9")
_OTM3_FILL = PatternFill("solid", fgColor="B39DDB")
RELATION_FILLS = {
    # Per-level (Phase 5B+)
    "ITM1": _ITM_FILL,
    "ITM2": _ITM2_FILL,
    "ITM3": _ITM3_FILL,
    "ATM":  _ATM_FILL,
    "OTM1": _OTM_FILL,
    "OTM2": _OTM2_FILL,
    "OTM3": _OTM3_FILL,
    # Legacy (pre-per-level signals.jsonl rows, before migration ran)
    "ITM":  _ITM_FILL,
    "OTM":  _OTM_FILL,
}

# would_alert_extended highlight
EXTENDED_FILL = PatternFill("solid", fgColor="FFE0B2")
ALL_PASSED_FILL = PatternFill("solid", fgColor="FFF59D")

# KPI tile palette (Strategy Dashboard)
KPI_FILLS = [
    PatternFill("solid", fgColor="1F4E78"),
    PatternFill("solid", fgColor="2E7D32"),
    PatternFill("solid", fgColor="EF6C00"),
    PatternFill("solid", fgColor="6A1B9A"),
    PatternFill("solid", fgColor="00838F"),
    PatternFill("solid", fgColor="C62828"),
]


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------


def _resolve_excel_path(year: int, quarter: int) -> Path:
    return DASHBOARDS_DIR / f"dashboard_{year:04d}_Q{quarter}.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _set_headers(ws: Worksheet, headers: Iterable[str], row: int = 1) -> None:
    for col_idx, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=header)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[row].height = 26


def _autofit(ws: Worksheet, df: pd.DataFrame, header_row: int = 1) -> None:
    """Approximate-autofit column widths from frame contents."""
    for col_idx, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].head(200).tolist()]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(12, max_len + 2), 48
        )


def _write_dataframe(
    ws: Worksheet,
    df: pd.DataFrame,
    header_row: int = 1,
    start_row: int | None = None,
) -> int:
    """Write ``df`` to ``ws``. Returns the row count written (excluding header).

    Uses Title-Case header strings derived from DataFrame columns. The
    raw column names are kept so callers know which cells to colour.
    """
    if df.empty:
        ws.cell(row=header_row, column=1, value="(no data yet)").font = _SUBTITLE_FONT
        return 0
    _set_headers(ws, [_pretty(c) for c in df.columns], row=header_row)
    sr = start_row if start_row is not None else header_row + 1
    for row_offset, (_, row) in enumerate(df.iterrows()):
        for col_idx, col in enumerate(df.columns, start=1):
            value = row[col]
            if pd.isna(value):
                value = None
            ws.cell(row=sr + row_offset, column=col_idx, value=value).font = _BODY_FONT
    _autofit(ws, df, header_row=header_row)
    return len(df)


def _pretty(col: str) -> str:
    """Convert ``opt_above_vwap_pct`` → ``Opt Above VWAP %``."""
    if col == "opt_above_vwap_pct":
        return "Opt Above VWAP %"
    if col == "bot_remark":
        return "Bot Remark"
    if col == "bot_tags":
        return "Bot Tags"
    if col == "outcome_remark":
        return "Outcome Remark"
    if col == "user_notes":
        return "User Notes"
    if col == "order_status":
        return "Order Status"
    if col == "exit_price":
        return "Exit Price"
    if col == "pnl_rupees":
        return "P&L"
    if col == "timestamp_ist":
        return "Timestamp IST"
    if col == "auto_order_status":
        return "Auto Order Status"
    if col == "auto_exit_price":
        return "Auto Exit Price"
    if col == "auto_exit_time":
        return "Auto Exit Time"
    if col == "auto_exit_reason":
        return "Auto Exit Reason"
    if col == "auto_pnl_per_unit":
        return "Auto P&L / unit"
    if col == "mfe":
        return "MFE"
    if col == "mae":
        return "MAE"
    if col == "intrabar_ambiguous":
        return "Intrabar Ambiguous"
    return col.replace("_", " ").title()


def _outcome_fill_for(value) -> PatternFill | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return OUTCOME_FILLS.get(str(value).strip().upper())


def _gap_fill_for(value) -> PatternFill | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return GAP_FILLS.get(str(value).strip().upper())


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_strategy_dashboard(
    ws: Worksheet, df: pd.DataFrame, year: int, quarter: int
) -> None:
    """KPI tiles + 4 charts. Designed to be the first thing the user sees."""
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1, value=f"SCC Strategy Dashboard — {year} Q{quarter}")
    title.font = _TITLE_FONT
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    subtitle = ws.cell(
        row=2,
        column=1,
        value=(
            "Auto-generated from Parquet · refreshes at 15:35 IST · "
            "filter the All Alerts / All Signals sheets for detail"
        ),
    )
    subtitle.font = _SUBTITLE_FONT

    alerts = df[df["event_type"] == "alert"] if "event_type" in df.columns else df.iloc[0:0]
    signals = df[df["event_type"].isin(["scan", "alert"])] if "event_type" in df.columns else df.iloc[0:0]
    extended = df[df["event_type"] == "would_alert_extended"] if "event_type" in df.columns else df.iloc[0:0]

    total_alerts = len(alerts)
    nifty_alerts = int(((alerts.get("symbol") == "NIFTY").sum()) if not alerts.empty else 0)
    bn_alerts = int(((alerts.get("symbol") == "BANKNIFTY").sum()) if not alerts.empty else 0)
    total_scans = len(signals)
    extended_count = len(extended)

    outcome_series = alerts.get("order_status") if "order_status" in alerts.columns else pd.Series(dtype=object)
    tp2 = int(((outcome_series == "TP2_HIT").sum()) if not alerts.empty else 0)
    tp1 = int(((outcome_series == "TP1_HIT").sum()) if not alerts.empty else 0)
    sl = int(((outcome_series == "SL_HIT").sum()) if not alerts.empty else 0)
    realised_pnl = float(alerts.get("pnl_rupees", pd.Series(dtype=float)).sum()) if not alerts.empty else 0.0
    filled_outcomes = int(outcome_series.notna().sum()) if not alerts.empty else 0
    win_rate = (
        (tp1 + tp2) / filled_outcomes * 100.0 if filled_outcomes > 0 else 0.0
    )

    kpi_rows = [
        ("Total Alerts", total_alerts, 0),
        ("NIFTY Alerts", nifty_alerts, 1),
        ("BankNifty Alerts", bn_alerts, 2),
        ("Total Scans", total_scans, 3),
        ("Extended-Zone Scans", extended_count, 4),
        ("Win Rate (filled)", f"{win_rate:.1f}%", 5),
        ("TP2 Hit", tp2, 0),
        ("TP1 Hit", tp1, 1),
        ("SL Hit", sl, 2),
        ("Filled Outcomes", filled_outcomes, 3),
        ("Realised P&L (₹)", f"{realised_pnl:,.0f}", 4),
        ("As of", datetime.now(IST).strftime("%Y-%m-%d %H:%M"), 5),
    ]

    base_row = 4
    for idx, (label, value, palette_idx) in enumerate(kpi_rows):
        row_off = (idx // 3) * 4
        col_off = (idx % 3) * 3
        top_row = base_row + row_off
        anchor_col = 1 + col_off

        # Top stripe — merged label centered across both tile columns.
        ws.merge_cells(
            start_row=top_row, start_column=anchor_col,
            end_row=top_row, end_column=anchor_col + 1,
        )
        label_cell = ws.cell(row=top_row, column=anchor_col, value=label.upper())
        label_cell.fill = KPI_FILLS[palette_idx % len(KPI_FILLS)]
        label_cell.font = Font(color="FFFFFF", bold=True, size=9)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[top_row].height = 20

        # Big value cell (merged 2-wide).
        ws.merge_cells(
            start_row=top_row + 1,
            start_column=anchor_col,
            end_row=top_row + 1,
            end_column=anchor_col + 1,
        )
        v_cell = ws.cell(row=top_row + 1, column=anchor_col, value=value)
        v_cell.font = Font(size=18, bold=True, color="1F4E78")
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        v_cell.border = _BORDER
        ws.row_dimensions[top_row + 1].height = 32

    # Set column widths for KPI grid.
    for col_letter in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col_letter].width = 18

    # ---------- charts ----------

    chart_start_row = base_row + (((len(kpi_rows) - 1) // 3) + 1) * 4 + 2

    # Helper: write a small aggregation table somewhere off-screen for chart refs.
    def _write_agg_table(
        title: str, items: list[tuple[str, float]], col_anchor: int
    ) -> tuple[int, int, int]:
        ws.cell(row=chart_start_row, column=col_anchor, value=title).font = Font(
            bold=True, color="1F4E78"
        )
        ws.cell(row=chart_start_row + 1, column=col_anchor, value="Bucket").font = _HEADER_FONT
        ws.cell(row=chart_start_row + 1, column=col_anchor).fill = _HEADER_FILL
        ws.cell(row=chart_start_row + 1, column=col_anchor + 1, value="Count").font = _HEADER_FONT
        ws.cell(row=chart_start_row + 1, column=col_anchor + 1).fill = _HEADER_FILL
        for i, (k, v) in enumerate(items, start=1):
            ws.cell(row=chart_start_row + 1 + i, column=col_anchor, value=k)
            ws.cell(row=chart_start_row + 1 + i, column=col_anchor + 1, value=v)
        return chart_start_row + 1, chart_start_row + 1 + len(items), col_anchor

    # 1. Wins vs Losses by Strike Relation
    # Relation labels are per-level (ITM3..ATM..OTM3) since Phase 5B+. The
    # display order keeps deeper ITM at the left and deeper OTM at the right;
    # any unexpected label (legacy or new) is appended at the end.
    _RELATION_DISPLAY_ORDER = (
        "ITM3", "ITM2", "ITM1", "ITM",  # "ITM" = legacy unmigrated
        "ATM",
        "OTM1", "OTM", "OTM2", "OTM3",  # "OTM" = legacy unmigrated
    )
    relation_alerts = alerts[alerts.get("relation").notna()] if not alerts.empty and "relation" in alerts.columns else pd.DataFrame()
    rel_buckets: list[tuple[str, float]] = []
    if not relation_alerts.empty:
        present = [r for r in _RELATION_DISPLAY_ORDER if (relation_alerts["relation"] == r).any()]
        # Tack on any unknown labels at the end so nothing is silently dropped.
        for extra in sorted(set(relation_alerts["relation"].dropna().astype(str)) - set(_RELATION_DISPLAY_ORDER)):
            present.append(extra)
        for relation in present:
            slice_ = relation_alerts[relation_alerts["relation"] == relation]
            wins = int(((slice_.get("order_status") == "TP2_HIT").sum() if "order_status" in slice_.columns else 0)
                       + ((slice_.get("order_status") == "TP1_HIT").sum() if "order_status" in slice_.columns else 0))
            losses = int(((slice_.get("order_status") == "SL_HIT").sum()) if "order_status" in slice_.columns else 0)
            rel_buckets.append((f"{relation} Win", wins))
            rel_buckets.append((f"{relation} Loss", losses))
    if not rel_buckets:
        rel_buckets = [("No filled outcomes yet", 0)]

    hdr_row, end_row, anchor = _write_agg_table(
        "Wins / Losses by Strike Relation", rel_buckets, col_anchor=10
    )
    if any(v > 0 for _, v in rel_buckets):
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 11
        chart1.title = "Wins vs Losses by Strike Relation"
        chart1.y_axis.title = "Count"
        chart1.x_axis.title = "Bucket"
        data_ref = Reference(ws, min_col=anchor + 1, min_row=hdr_row, max_row=end_row, max_col=anchor + 1)
        cats_ref = Reference(ws, min_col=anchor, min_row=hdr_row + 1, max_row=end_row)
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)
        chart1.dataLabels = DataLabelList(showVal=True)
        ws.add_chart(chart1, f"A{chart_start_row}")

    # 2. Alerts by VIX Regime
    if not alerts.empty and "vix_regime" in alerts.columns:
        vix_counts = alerts["vix_regime"].fillna("Unknown").value_counts().to_dict()
    else:
        vix_counts = {}
    vix_items = list(vix_counts.items()) or [("(no alerts)", 0)]
    hdr2, end2, anc2 = _write_agg_table(
        "Alerts by VIX Regime", vix_items, col_anchor=13
    )
    if any(v > 0 for _, v in vix_items):
        chart2 = BarChart()
        chart2.type = "bar"
        chart2.style = 12
        chart2.title = "Alerts by VIX Regime"
        data_ref = Reference(ws, min_col=anc2 + 1, min_row=hdr2, max_row=end2, max_col=anc2 + 1)
        cats_ref = Reference(ws, min_col=anc2, min_row=hdr2 + 1, max_row=end2)
        chart2.add_data(data_ref, titles_from_data=True)
        chart2.set_categories(cats_ref)
        chart2.dataLabels = DataLabelList(showVal=True)
        ws.add_chart(chart2, f"E{chart_start_row}")

    # 3. Alerts by Time of Day (30-min buckets)
    time_buckets: dict[str, int] = {}
    if not alerts.empty and "time" in alerts.columns:
        for t in alerts["time"].dropna().astype(str):
            try:
                hh, mm = int(t[:2]), int(t[3:5])
            except Exception:
                continue
            anchor_minute = (mm // 30) * 30
            key = f"{hh:02d}:{anchor_minute:02d}"
            time_buckets[key] = time_buckets.get(key, 0) + 1
    time_items = sorted(time_buckets.items()) or [("(no alerts)", 0)]
    hdr3, end3, anc3 = _write_agg_table(
        "Alerts by 30-min Time Bucket", time_items, col_anchor=16
    )
    if any(v > 0 for _, v in time_items):
        chart3 = BarChart()
        chart3.type = "col"
        chart3.style = 13
        chart3.title = "Alerts by Time of Day (30-min buckets)"
        chart3.x_axis.title = "Time bucket"
        chart3.y_axis.title = "Alerts"
        data_ref = Reference(ws, min_col=anc3 + 1, min_row=hdr3, max_row=end3, max_col=anc3 + 1)
        cats_ref = Reference(ws, min_col=anc3, min_row=hdr3 + 1, max_row=end3)
        chart3.add_data(data_ref, titles_from_data=True)
        chart3.set_categories(cats_ref)
        chart3.dataLabels = DataLabelList(showVal=True)
        ws.add_chart(chart3, f"A{chart_start_row + 16}")

    # 4. Cumulative P&L line
    cum_items: list[tuple[str, float]] = []
    if not alerts.empty and "pnl_rupees" in alerts.columns:
        pnl_rows = alerts[alerts["pnl_rupees"].notna()].sort_values("timestamp_ist")
        running = 0.0
        for _, r in pnl_rows.iterrows():
            running += float(r["pnl_rupees"])
            label = str(r.get("date") or r.get("timestamp_ist", ""))[:10]
            cum_items.append((label, running))
    if not cum_items:
        cum_items = [("(no filled P&L)", 0.0)]
    hdr4, end4, anc4 = _write_agg_table(
        "Cumulative P&L", cum_items, col_anchor=19
    )
    if any(isinstance(v, (int, float)) and v != 0 for _, v in cum_items):
        chart4 = LineChart()
        chart4.style = 12
        chart4.title = "Cumulative P&L (₹)"
        chart4.x_axis.title = "Date"
        chart4.y_axis.title = "Cumulative P&L"
        data_ref = Reference(ws, min_col=anc4 + 1, min_row=hdr4, max_row=end4, max_col=anc4 + 1)
        cats_ref = Reference(ws, min_col=anc4, min_row=hdr4 + 1, max_row=end4)
        chart4.add_data(data_ref, titles_from_data=True)
        chart4.set_categories(cats_ref)
        ws.add_chart(chart4, f"E{chart_start_row + 16}")


def _build_daily_summary(ws: Worksheet, df: pd.DataFrame) -> int:
    if df.empty or "date" not in df.columns:
        _set_headers(ws, ["Date", "Note"])
        ws.cell(row=2, column=1, value="(no data)").font = _SUBTITLE_FONT
        return 0

    alerts = df[df["event_type"] == "alert"]
    signals = df[df["event_type"].isin(["scan", "alert"])]
    extended = df[df["event_type"] == "would_alert_extended"]
    rejections = df[df["event_type"] == "rejection"]
    gaps = df[df["event_type"] == "gap"]

    dates = sorted({d for d in df["date"].dropna().unique()})

    rows: list[dict] = []
    for d in dates:
        gap_row = gaps[gaps["date"] == d]
        decision = (
            gap_row.iloc[0]["decision"] if not gap_row.empty and "decision" in gap_row.columns else "NORMAL"
        )
        nifty_gap = (
            gap_row.iloc[0].get("nifty_gap_pct") if not gap_row.empty else None
        )
        bn_gap = (
            gap_row.iloc[0].get("banknifty_gap_pct") if not gap_row.empty else None
        )
        rows.append({
            "date": d,
            "gap_decision": decision,
            "nifty_gap_pct": nifty_gap,
            "banknifty_gap_pct": bn_gap,
            "scans": int(len(signals[signals["date"] == d])),
            "alerts": int(len(alerts[alerts["date"] == d])),
            "extended": int(len(extended[extended["date"] == d])),
            "rejections": int(len(rejections[rejections["date"] == d])),
            "nifty_alerts": int(((alerts["date"] == d) & (alerts.get("symbol") == "NIFTY")).sum())
                if not alerts.empty else 0,
            "banknifty_alerts": int(((alerts["date"] == d) & (alerts.get("symbol") == "BANKNIFTY")).sum())
                if not alerts.empty else 0,
        })

    out = pd.DataFrame(rows)
    written = _write_dataframe(ws, out)

    # Colour the Gap Decision cell.
    if "gap_decision" in out.columns and written:
        gap_col_idx = list(out.columns).index("gap_decision") + 1
        for i in range(written):
            cell = ws.cell(row=2 + i, column=gap_col_idx)
            fill = _gap_fill_for(cell.value)
            if fill:
                cell.fill = fill
    ws.freeze_panes = "A2"
    return written


def _build_all_alerts(ws: Worksheet, df: pd.DataFrame) -> int:
    alerts = df[df["event_type"] == "alert"] if "event_type" in df.columns else pd.DataFrame()
    if alerts.empty:
        _set_headers(ws, ["(no alerts yet)"])
        return 0

    show_cols = [
        "timestamp_ist", "date", "time", "symbol", "strike", "relation",
        "option_type", "expiry", "spot_price", "entry", "sl", "tp1", "tp2",
        "lots", "total_risk", "vix", "vix_regime", "day_type",
        "opt_above_vwap_pct", "rsi", "rsi_ma", "oi", "oi_ma",
        "volume", "volume_ma",
        "bot_remark", "bot_tags",
    ]
    cols = [c for c in show_cols if c in alerts.columns]
    out = alerts[cols].sort_values("timestamp_ist").reset_index(drop=True)
    written = _write_dataframe(ws, out)

    # Wrap text on bot_remark / bot_tags columns.
    for wrap_col in ("bot_remark", "bot_tags"):
        if wrap_col in out.columns:
            ci = list(out.columns).index(wrap_col) + 1
            ws.column_dimensions[get_column_letter(ci)].width = 60
            for r in range(2, written + 2):
                ws.cell(row=r, column=ci).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
    ws.freeze_panes = "A2"
    return written


def _build_order_place(ws: Worksheet, df: pd.DataFrame) -> int:
    alerts = df[df["event_type"] == "alert"] if "event_type" in df.columns else pd.DataFrame()

    auto_cols = [
        "timestamp_ist", "date", "time", "symbol", "strike", "relation",
        "option_type", "entry", "sl", "tp1", "tp2", "lots", "total_risk",
        "bot_remark",
    ]
    manual_cols = ["order_status", "exit_price", "pnl_rupees", "outcome_remark", "user_notes"]
    # Phase 5B-A: virtual-replay columns. Appended after the manual block
    # so the manual cells stay in the user's expected position.
    auto_outcome_cols = [
        "auto_order_status", "auto_exit_price", "auto_exit_time",
        "auto_exit_reason", "auto_pnl_per_unit",
        "mfe", "mae", "intrabar_ambiguous",
    ]

    if alerts.empty:
        _set_headers(ws, [_pretty(c) for c in auto_cols + manual_cols + auto_outcome_cols])
        ws.cell(row=2, column=1, value="(no alerts yet — manual columns activate after first alert)").font = _SUBTITLE_FONT
        return 0

    out = alerts.copy()
    for col in manual_cols + auto_outcome_cols:
        if col not in out.columns:
            out[col] = None
    out = out[auto_cols + manual_cols + auto_outcome_cols].sort_values("timestamp_ist").reset_index(drop=True)
    written = _write_dataframe(ws, out)

    # Colour Order Status cell.
    if "order_status" in out.columns and written:
        status_col_idx = list(out.columns).index("order_status") + 1
        for i in range(written):
            cell = ws.cell(row=2 + i, column=status_col_idx)
            fill = _outcome_fill_for(cell.value)
            if fill:
                cell.fill = fill

    # Colour auto_order_status cell (same palette as manual).
    if "auto_order_status" in out.columns and written:
        auto_idx = list(out.columns).index("auto_order_status") + 1
        for i in range(written):
            cell = ws.cell(row=2 + i, column=auto_idx)
            fill = _outcome_fill_for(cell.value)
            if fill:
                cell.fill = fill

    # Colour P&L cell green/red.
    if "pnl_rupees" in out.columns and written:
        pnl_idx = list(out.columns).index("pnl_rupees") + 1
        green = PatternFill("solid", fgColor="C8E6C9")
        red = PatternFill("solid", fgColor="EF9A9A")
        for i in range(written):
            cell = ws.cell(row=2 + i, column=pnl_idx)
            v = cell.value
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            cell.fill = green if fv >= 0 else red

    # Wrap text on remark columns.
    for wrap_col in ("bot_remark", "outcome_remark", "user_notes"):
        if wrap_col in out.columns:
            ci = list(out.columns).index(wrap_col) + 1
            ws.column_dimensions[get_column_letter(ci)].width = 48
            for r in range(2, written + 2):
                ws.cell(row=r, column=ci).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    # Legend strip below the data.
    legend_row = written + 3
    ws.cell(row=legend_row, column=1, value="Outcome legend:").font = Font(bold=True)
    for i, (label, fill) in enumerate(OUTCOME_FILLS.items(), start=2):
        cell = ws.cell(row=legend_row, column=i, value=label)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    return written


def _build_all_signals(ws: Worksheet, df: pd.DataFrame) -> int:
    if df.empty or "event_type" not in df.columns:
        _set_headers(ws, ["(no signals yet)"])
        return 0

    signals = df[df["event_type"].isin(["scan", "alert", "would_alert_extended"])].copy()
    if signals.empty:
        _set_headers(ws, ["(no signals yet)"])
        return 0

    show_cols = [
        "timestamp_ist", "date", "time", "event_type", "symbol", "strike",
        "relation", "option_type", "spot_price", "spot_vwap",
        "option_close", "option_vwap", "opt_above_vwap_pct",
        "rsi", "rsi_ma", "oi", "oi_ma", "volume", "volume_ma",
        "is_green", "all_passed", "summary",
    ]
    cols = [c for c in show_cols if c in signals.columns]
    if "time" not in signals.columns and "timestamp_ist" in signals.columns:
        signals["time"] = signals["timestamp_ist"].astype(str).str.slice(11, 16)
        cols = [c for c in show_cols if c in signals.columns]

    out = signals[cols].sort_values("timestamp_ist").reset_index(drop=True)
    written = _write_dataframe(ws, out)

    # Highlight all_passed → yellow; would_alert_extended → light orange.
    if written:
        all_passed_col = (
            list(out.columns).index("all_passed") + 1 if "all_passed" in out.columns else None
        )
        event_col = list(out.columns).index("event_type") + 1
        relation_col = (
            list(out.columns).index("relation") + 1 if "relation" in out.columns else None
        )
        for i in range(written):
            event_val = ws.cell(row=2 + i, column=event_col).value
            if event_val == "would_alert_extended":
                for c in range(1, len(out.columns) + 1):
                    ws.cell(row=2 + i, column=c).fill = EXTENDED_FILL
                continue
            ap_val = ws.cell(row=2 + i, column=all_passed_col).value if all_passed_col else False
            if ap_val is True or ap_val == "True":
                for c in range(1, len(out.columns) + 1):
                    ws.cell(row=2 + i, column=c).fill = ALL_PASSED_FILL
                continue
            if relation_col is not None:
                rel = ws.cell(row=2 + i, column=relation_col).value
                fill = RELATION_FILLS.get(str(rel)) if rel else None
                if fill:
                    for c in range(1, len(out.columns) + 1):
                        ws.cell(row=2 + i, column=c).fill = fill
    ws.freeze_panes = "A2"
    return written



def _build_gap_history(ws: Worksheet, df: pd.DataFrame) -> int:
    if df.empty or "event_type" not in df.columns:
        _set_headers(ws, ["(no gap history)"])
        return 0
    gaps = df[df["event_type"] == "gap"].copy()
    if gaps.empty:
        _set_headers(ws, ["(no gap history)"])
        return 0
    cols = [
        c for c in (
            "timestamp_ist", "date", "decision", "enabled", "threshold_pct",
            "direction", "any_triggered",
            "nifty_open", "nifty_prev_close", "nifty_gap_pct",
            "banknifty_open", "banknifty_prev_close", "banknifty_gap_pct",
        ) if c in gaps.columns
    ]
    out = gaps[cols].sort_values("timestamp_ist").reset_index(drop=True)
    written = _write_dataframe(ws, out)

    # Colour the Decision cell.
    if "decision" in out.columns and written:
        dec_col_idx = list(out.columns).index("decision") + 1
        for i in range(written):
            cell = ws.cell(row=2 + i, column=dec_col_idx)
            fill = _gap_fill_for(cell.value)
            if fill:
                cell.fill = fill
    ws.freeze_panes = "A2"
    return written


def _build_config_snapshot(ws: Worksheet) -> None:
    from src.config_loader import load_config

    project_root = Path(__file__).resolve().parents[2]
    config_path = project_root / "config" / "config.yaml"
    try:
        cfg = load_config(config_path)
    except Exception as e:
        ws.cell(row=1, column=1, value=f"Failed to load config: {e}").font = _SUBTITLE_FONT
        return

    rows = []
    rows.append(("Generated at", datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")))
    rows.append(("Active feed", cfg.feeds.active_feed))
    rows.append(("Alert mode", cfg.mode.alert_mode))
    rows.append(("Order place mode", cfg.mode.order_place_mode))
    rows.append(("Paper trade mode", cfg.mode.paper_trade_mode))
    rows.append(("NIFTY enabled", cfg.instruments.nifty_enabled))
    rows.append(("BankNifty enabled", cfg.instruments.banknifty_enabled))
    rows.append(("NIFTY lot size", cfg.instruments.nifty_lot_size))
    rows.append(("BankNifty lot size", cfg.instruments.banknifty_lot_size))
    rows.append(("Normal start time", cfg.time_rules.normal_start_time))
    rows.append(("Gap-day start time", cfg.time_rules.gap_day_start_time))
    rows.append(("Last entry time", cfg.time_rules.last_entry_time))
    rows.append(("Hard squareoff time", cfg.time_rules.hard_squareoff_time))
    rows.append(("Gap-day enabled", cfg.time_rules.gap_day_enabled))
    rows.append(("Gap-day threshold %", cfg.time_rules.gap_day_threshold_pct))
    rows.append(("Gap-day direction", cfg.time_rules.gap_day_direction))
    rows.append(("Target risk / trade (₹)", cfg.risk_reward.target_risk_per_trade))
    rows.append(("Risk range", f"₹{cfg.risk_reward.risk_range_min}–{cfg.risk_reward.risk_range_max}"))
    rows.append(("TP1 / TP2 (normal)", f"{cfg.risk_reward.normal_day_tp1_r}R / {cfg.risk_reward.normal_day_tp2_r}R"))
    rows.append(("TP1 / TP2 (expiry)", f"{cfg.risk_reward.expiry_day_tp1_r}R / {cfg.risk_reward.expiry_day_tp2_r}R"))
    rows.append(("Max SL per day", cfg.circuit_breakers.max_sl_per_day))
    rows.append(("Max loss per day (₹)", cfg.circuit_breakers.max_loss_per_day_rupees))
    rows.append(("RSI min / max", f"{cfg.conditions.c3_rsi_min} – {cfg.conditions.c3_rsi_max}"))
    rows.append(("C1 max distance %", cfg.conditions.c1_max_distance_pct))
    rows.append(("C1 extended zone max %", cfg.conditions.c1_extended_zone_max_pct))
    rows.append(("Auto-dashboard at 15:35", cfg.dashboard.auto_trigger_at_1535))

    _set_headers(ws, ["Setting", "Value"])
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = _BODY_FONT
        ws.cell(row=i, column=2, value=v).font = _BODY_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30


# ---------------------------------------------------------------------------
# Public — update_dashboard
# ---------------------------------------------------------------------------


def update_dashboard(feed: "Any | None" = None) -> dict:
    """Refresh every quarter touched by the data files.

    Args:
        feed: Optional active ``BaseFeed`` instance. When provided (and
            the trading day is complete), ``_append_paper_sheets`` will
            fetch candles via the feed for any replay-cache miss so that
            paper-trade outcomes are computed instead of NO_DATA.
            Pass ``None`` for cache-only mode (manual mid-day runs).

    Returns a dict with status info. The latest-quarter path is in
    ``output_path``. Idempotent.
    """
    DASHBOARDS_DIR.mkdir(parents=True, exist_ok=True)

    now_ist = datetime.now(IST)
    today_q = quarter_for_date(now_ist)

    # Find every (year, quarter) present in the Parquet data.
    from src.dashboard.data_writer import _all_parquet_months  # local import to avoid cycle

    months = _all_parquet_months()
    quarters: set[tuple[int, int]] = set()
    for p in months:
        try:
            month_str = p.stem.replace("scc_data_", "")
            year, month = map(int, month_str.split("-"))
            quarters.add((year, (month - 1) // 3 + 1))
        except Exception:
            continue

    # Always include the current quarter so the file exists for live runs
    # even before any data has accumulated.
    quarters.add(today_q)

    if not quarters:
        return {"status": "no_data", "output_path": None}

    latest_written: Path | None = None
    counts = {
        "alerts_added": 0,
        "signals_added": 0,
        "order_place_added": 0,
        "gaps_added": 0,
        "quarters_touched": 0,
    }

    for year, quarter in sorted(quarters):
        df = load_parquet_for_quarter(year, quarter)
        path = _build_workbook(year, quarter, df, feed=feed)
        latest_written = path
        counts["quarters_touched"] += 1
        if not df.empty and "event_type" in df.columns:
            counts["alerts_added"] += int((df["event_type"] == "alert").sum())
            counts["signals_added"] += int(
                df["event_type"].isin(["scan", "alert", "would_alert_extended"]).sum()
            )
            counts["order_place_added"] += int((df["event_type"] == "alert").sum())
            counts["gaps_added"] += int((df["event_type"] == "gap").sum())

    return {
        "status": "ok",
        "output_path": str(latest_written) if latest_written else None,
        **counts,
    }


def _build_workbook(year: int, quarter: int, df: pd.DataFrame, feed: "Any | None" = None) -> Path:
    wb = Workbook()
    # Default sheet → Strategy Dashboard.
    dash = wb.active
    dash.title = "Strategy Dashboard"
    _build_strategy_dashboard(dash, df, year, quarter)

    daily = wb.create_sheet("Daily Summary")
    _build_daily_summary(daily, df)

    alerts = wb.create_sheet("All Alerts")
    _build_all_alerts(alerts, df)

    order = wb.create_sheet("Order Place")
    _build_order_place(order, df)

    signals = wb.create_sheet("All Signals")
    _build_all_signals(signals, df)

    gaps = wb.create_sheet("Gap History")
    _build_gap_history(gaps, df)

    # Phase 5D — paper-trade sheets. Best-effort: if the engine fails
    # (no alerts file, mid-day run, etc.) the rest of the workbook is
    # untouched. Sheets are appended AFTER the existing ones so the
    # default-open sheet stays Strategy Dashboard.
    _append_paper_sheets(wb, feed=feed)

    cfg_sheet = wb.create_sheet("Config Snapshot")
    _build_config_snapshot(cfg_sheet)

    path = _resolve_excel_path(year, quarter)
    wb.save(path)
    logger.info(f"Dashboard written: {path}")
    return path


def _append_paper_sheets(wb: "Workbook", feed: "Any | None" = None) -> None:
    """Append the Phase 5D Paper Trades / Dashboard / Echoes sheets.

    Best-effort: any failure inside the paper layer is logged at
    warning and the workbook continues. Existing sheets are never
    touched here.

    Args:
        feed: Optional live ``BaseFeed``. When provided the candle cache
            is populated for any completed trading day that is missing
            cached candles, enabling outcome replay. ``None`` → cache-only
            (outcomes already in cache replay fine; uncached days stay
            NO_DATA until the next feed-connected sync).
    """
    try:
        from src.config_loader import load_config
        from src.paper.engine import run_paper_engine
        from src.paper.persistence import (
            ensure_overrides_file,
            merge_overrides,
            read_overrides,
            read_paper_trades,
        )
        from src.dashboard.paper_sheets import (
            build_echoes_sheet,
            build_paper_dashboard_sheet,
            build_paper_trades_sheet,
        )
        from src.dashboard.candle_cache import get_or_fetch_candles

        project_root = Path(__file__).resolve().parents[2]
        cfg = load_config(project_root / "config" / "config.yaml")
        if not cfg.paper_trading.enabled:
            return

        def _candle_source(symbol, strike, option_type, expiry, trading_date):
            return get_or_fetch_candles(
                feed=feed,
                symbol=symbol,
                strike=strike,
                option_type=option_type,
                expiry=expiry,
                trading_date=trading_date,
            )

        alerts_path = project_root / "logs" / "alerts.jsonl"
        result = run_paper_engine(
            alerts_path=str(alerts_path),
            app_config=cfg,
            candle_source=_candle_source,
            write=True,
            compute_all_alerts=False,
        )

        ensure_overrides_file(result.overrides_path)
        trades_df = read_paper_trades(result.paper_trades_path)
        overrides_df = read_overrides(result.overrides_path)
        merged = merge_overrides(trades_df, overrides_df)

        paper_ws = wb.create_sheet("Paper Trades")
        build_paper_trades_sheet(paper_ws, merged)

        dash_ws = wb.create_sheet("Paper Dashboard")
        build_paper_dashboard_sheet(
            dash_ws, merged, collapse_summary=result.collapse_summary
        )

        echo_ws = wb.create_sheet("Echoes (diagnostic)")
        build_echoes_sheet(echo_ws, result.annotated_alerts)
    except Exception as e:
        logger.warning(f"paper sheets: skipped — {e}")
